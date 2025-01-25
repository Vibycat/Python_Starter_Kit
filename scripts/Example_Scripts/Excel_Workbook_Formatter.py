"""
**********************************************
*            Script Information             *
**********************************************
Script Name: Excel Workbook Formatter
Author: [Kyle May]
Date: [1-12-2025]
Description:
    Demonstrates various features of the openpyxl library to work with Excel files.
    Includes examples of adding headers, formatting cells, adding formulas, and more.
**********************************************
"""

# **********************************************
# *             Import Libraries              *
# **********************************************
# The `openpyxl` library is used for creating and manipulating Excel files (.xlsx).
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
# The `os` library is used for handling file paths and directory operations.
import os

# **********************************************
# *           Function Definitions            *
# **********************************************
def create_workbook_with_headers(wb):
    """
    Adds headers to the default worksheet and applies formatting.
    Args:
        wb (Workbook): The openpyxl workbook instance.
    """
    ws = wb.active
    ws.title = "Employee Info"
    headers = ["Name", "Age", "Gender", "Race", "Title", "Favorite Food", "Pets"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)  # Make the header text bold for emphasis.
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align the header text.
    ws.freeze_panes = "A2"  # Freeze the top row so headers remain visible when scrolling.

def add_sample_data(ws):
    """
    Adds sample data to the worksheet.
    Args:
        ws (Worksheet): The openpyxl worksheet instance.
    """
    # Example data to populate the worksheet.
    data = [
        ["John", 30, "Male", "Human", "Engineer", "Pizza", "Dog"],
        ["Sara", 25, "Female", "Elf", "Designer", "Sushi", "Cat"],
        ["Mike", 40, "Male", "Orc", "Manager", "Steak", "None"],
    ]
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

def adjust_column_widths(ws):
    """
    Adjusts column widths based on the content.
    Args:
        ws (Worksheet): The openpyxl worksheet instance.
    """
    # Dynamically adjust each column's width based on the longest entry.
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column's letter representation (e.g., A, B).
        for cell in col:
            if cell.value:  # Only consider cells with values.
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding for better readability.

def add_formatting_examples(ws):
    """
    Demonstrates various cell formatting options.
    Args:
        ws (Worksheet): The openpyxl worksheet instance.
    """
    # Merge cells and add a title.
    ws.merge_cells('A5:C5')
    ws['A5'] = "Merged Cells Example"

    # Adjust row height and column width.
    ws.row_dimensions[5].height = 20  # Set row height for better visibility.
    ws.column_dimensions['D'].width = 25  # Set a custom width for column D.

    # Style text with bold, italic, and color.
    ws['A6'] = "Styled Text"
    ws['A6'].font = Font(bold=True, italic=True, color="FF0000")

    # Apply a background color to a cell.
    ws['A7'] = "Colored Cell"
    ws['A7'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Center-align text in a cell.
    ws['B7'] = "Center Aligned"
    ws['B7'].alignment = Alignment(horizontal="center", vertical="center")

    # Add a formula to sum values in column B.
    ws['B4'] = "=SUM(B2:B3)"

    # Add a border to a cell.
    ws['C7'] = "Bordered Cell"
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws['C7'].border = thin_border

def save_workbook(wb, save_path):
    """
    Saves the workbook to the specified path.
    Args:
        wb (Workbook): The openpyxl workbook instance.
        save_path (str): Path where the workbook will be saved.
    """
    if not os.path.exists(save_path):
        os.makedirs(save_path)  # Create the directory if it doesn't exist.
    file_path = os.path.join(save_path, "MyNewWorkbook.xlsx")
    wb.save(file_path)
    print(f"Workbook saved successfully at {file_path}")

# **********************************************
# *                Main Execution             *
# **********************************************
if __name__ == "__main__":
    # Create a new workbook instance.
    wb = Workbook()
    ws = wb.active  # Get the active worksheet.

    # Execute the defined functions to create and format the workbook.
    create_workbook_with_headers(wb)
    add_sample_data(ws)
    adjust_column_widths(ws)
    add_formatting_examples(ws)

    # Define the path to save the workbook.
    save_path = r"G:\Excel\Workbooks\Testing_Workbooks"  # Update this path to a valid location on your system.
    save_workbook(wb, save_path)
